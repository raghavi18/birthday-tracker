"""
Team Birthday Tracker - Flask Backend
SQLite (local dev, no DATABASE_URL) or PostgreSQL (production, DATABASE_URL set).
"""
import os
import re
import csv
import io
import json
import sqlite3
from datetime import datetime, date
from contextlib import contextmanager

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DB_PATH     = os.environ.get("DB_PATH",    os.path.join(BASE_DIR, "birthdays.db"))
BACKUP_DIR  = os.environ.get("BACKUP_DIR", os.path.join(BASE_DIR, "backups"))
LATEST_BACKUP = os.path.join(BACKUP_DIR, "latest.json")

# Use PostgreSQL when DATABASE_URL is provided; fall back to SQLite for local dev.
DATABASE_URL = os.environ.get("DATABASE_URL", "")
_IS_PG = bool(DATABASE_URL)

if _IS_PG:
    # pg8000 is a pure-Python PostgreSQL driver — no compiled extensions, works on any Python version.
    import pg8000.dbapi as _pg8000
    from urllib.parse import urlparse as _urlparse
    import ssl as _ssl
else:
    os.makedirs(BACKUP_DIR, exist_ok=True)

# ORDER BY clause differs: PostgreSQL has no COLLATE NOCASE.
_NAME_ORDER = "LOWER(name)" if _IS_PG else "name COLLATE NOCASE"

app = Flask(__name__)

# Explicitly allow the GitHub Pages frontend (and localhost for local dev).
_FRONTEND_ORIGIN = os.environ.get("FRONTEND_URL", "https://raghavi18.github.io")
CORS(app, origins=[
    _FRONTEND_ORIGIN,
    "http://localhost:8000",
    "http://127.0.0.1:8000",
    "http://localhost:5050",
    "http://127.0.0.1:5050",
    "null",  # file:// origin when opening HTML files directly during local dev
])


# ---------------------------------------------------------------------------
# Database abstraction — identical call sites for SQLite and PostgreSQL.
# ---------------------------------------------------------------------------

def _pg_connect():
    """Open a pg8000 connection from DATABASE_URL, adding SSL for cloud hosts."""
    p = _urlparse(DATABASE_URL)
    kwargs = {
        "user": p.username,
        "password": p.password,
        "host": p.hostname,
        "port": p.port or 5432,
        "database": p.path.lstrip("/"),
    }
    if p.hostname and p.hostname not in ("localhost", "127.0.0.1"):
        # Supabase pooler uses a self-signed cert chain; disable verification while keeping encryption.
        ctx = _ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = _ssl.CERT_NONE
        kwargs["ssl_context"] = ctx
    return _pg8000.connect(**kwargs)


class _PGCursor:
    """Wraps a pg8000 cursor so fetchall/fetchone return dicts instead of tuples."""

    def __init__(self, cur):
        self._cur = cur

    def _cols(self):
        return [d[0] for d in (self._cur.description or [])]

    def fetchall(self):
        rows = self._cur.fetchall() or []
        cols = self._cols()
        return [dict(zip(cols, row)) for row in rows]

    def fetchone(self):
        row = self._cur.fetchone()
        if row is None:
            return None
        return dict(zip(self._cols(), row))


class _DB:
    """Thin wrapper normalizing SQLite and PostgreSQL for the same call pattern."""

    def __init__(self, raw_conn, is_pg: bool):
        self._conn = raw_conn
        self._is_pg = is_pg

    def execute(self, sql: str, params=()):
        """Run sql with params; return a cursor with .fetchall() / .fetchone()."""
        if self._is_pg:
            cur = self._conn.cursor()
            cur.execute(sql.replace("?", "%s"), params or None)
            return _PGCursor(cur)
        else:
            return self._conn.execute(sql, params or ())

    def insert_get_id(self, sql: str, params=()) -> int:
        """Execute INSERT and return the new row's primary key."""
        if self._is_pg:
            cur = self._conn.cursor()
            cur.execute((sql + " RETURNING id").replace("?", "%s"), params or None)
            row = cur.fetchone()
            return row[0]  # RETURNING id gives a single-element tuple
        else:
            cur = self._conn.execute(sql, params or ())
            return cur.lastrowid


@contextmanager
def get_db():
    if _IS_PG:
        conn = _pg_connect()
        db = _DB(conn, is_pg=True)
        try:
            yield db
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
    else:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        db = _DB(conn, is_pg=False)
        try:
            yield db
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()


# ---------------------------------------------------------------------------
# Schema SQL (auto-increment syntax differs between PostgreSQL and SQLite)
# ---------------------------------------------------------------------------

_SCHEMA_PG = """
    CREATE TABLE IF NOT EXISTS members (
        id             BIGSERIAL PRIMARY KEY,
        name           TEXT NOT NULL,
        role           TEXT NOT NULL,
        birthday_month INTEGER NOT NULL CHECK(birthday_month BETWEEN 1 AND 12),
        birthday_day   INTEGER NOT NULL CHECK(birthday_day   BETWEEN 1 AND 31),
        email          TEXT NOT NULL UNIQUE,
        created_at     TEXT NOT NULL,
        updated_at     TEXT NOT NULL,
        deleted_at     TEXT
    )
"""

_SCHEMA_SQLITE = """
    CREATE TABLE IF NOT EXISTS members (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        name           TEXT NOT NULL,
        role           TEXT NOT NULL,
        birthday_month INTEGER NOT NULL CHECK(birthday_month BETWEEN 1 AND 12),
        birthday_day   INTEGER NOT NULL CHECK(birthday_day   BETWEEN 1 AND 31),
        email          TEXT NOT NULL UNIQUE,
        created_at     TEXT NOT NULL,
        updated_at     TEXT NOT NULL,
        deleted_at     TEXT
    )
"""

_SCHEMA_SQLITE_BARE = """
    CREATE TABLE IF NOT EXISTS members (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        name           TEXT NOT NULL,
        role           TEXT NOT NULL,
        birthday_month INTEGER NOT NULL,
        birthday_day   INTEGER NOT NULL,
        email          TEXT NOT NULL UNIQUE,
        created_at     TEXT NOT NULL,
        updated_at     TEXT NOT NULL,
        deleted_at     TEXT
    )
"""


def init_db():
    """Initialize schema. SQLite-only: attempts backup restore on corruption."""
    try:
        with get_db() as db:
            db.execute(_SCHEMA_PG if _IS_PG else _SCHEMA_SQLITE)
    except Exception as e:
        if _IS_PG:
            app.logger.error(f"PostgreSQL init error: {e}")
            raise
        app.logger.warning(f"SQLite init error ({e}); attempting backup restore.")
        if os.path.exists(LATEST_BACKUP):
            restore_from_backup()
        else:
            if os.path.exists(DB_PATH):
                os.rename(DB_PATH, DB_PATH + ".corrupt")
            with get_db() as db:
                db.execute(_SCHEMA_SQLITE_BARE)


def restore_from_backup():
    """Rebuild SQLite DB from latest JSON backup (SQLite only; PG data persists natively)."""
    if _IS_PG:
        return
    if os.path.exists(DB_PATH):
        os.rename(DB_PATH, DB_PATH + ".corrupt")
    with open(LATEST_BACKUP, "r", encoding="utf-8") as f:
        data = json.load(f)
    with get_db() as db:
        db.execute(_SCHEMA_SQLITE)
        for row in data.get("members", []):
            db.execute(
                """INSERT INTO members (id, name, role, birthday_month, birthday_day,
                                        email, created_at, updated_at, deleted_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (row.get("id"), row["name"], row["role"],
                 row["birthday_month"], row["birthday_day"], row["email"],
                 row["created_at"], row["updated_at"], row.get("deleted_at")),
            )
    app.logger.info("SQLite database restored from backup.")


def write_backup():
    """Write timestamped JSON snapshot to local backups/ (SQLite only).
    Skipped in PostgreSQL mode — the DB itself is persistent on the cloud provider."""
    if _IS_PG:
        return
    with get_db() as db:
        rows = db.execute("SELECT * FROM members").fetchall()
    payload = {
        "exported_at": datetime.utcnow().isoformat() + "Z",
        "members": [dict(r) for r in rows],
    }
    ts = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    snapshot = os.path.join(BACKUP_DIR, f"members_{ts}.json")
    with open(snapshot, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    with open(LATEST_BACKUP, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------
EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")
NAME_RE  = re.compile(r"^[A-Za-z][A-Za-z\s\-']{1,59}$")

DAYS_IN_MONTH = {
    1: 31, 2: 29, 3: 31, 4: 30, 5: 31, 6: 30,
    7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31,
}


def validate_member_payload(p, *, partial=False):
    """Returns (cleaned_dict, error_dict). error_dict is empty if all good."""
    errors = {}
    cleaned = {}

    if "name" in p or not partial:
        name = (p.get("name") or "").strip()
        if not name:
            errors["name"] = "Name is required."
        elif len(name) < 2 or len(name) > 60:
            errors["name"] = "Name must be 2–60 characters."
        elif not NAME_RE.match(name):
            errors["name"] = "Name may only contain letters, spaces, hyphens, and apostrophes."
        else:
            cleaned["name"] = name

    if "role" in p or not partial:
        role = (p.get("role") or "").strip()
        if not role:
            errors["role"] = "Role is required."
        elif len(role) > 80:
            errors["role"] = "Role must be 80 characters or fewer."
        else:
            cleaned["role"] = role

    if "birthday_month" in p or "birthday_day" in p or not partial:
        try:
            m = int(p.get("birthday_month"))
            d = int(p.get("birthday_day"))
        except (TypeError, ValueError):
            errors["birthday"] = "Birthday month and day are required."
        else:
            if m < 1 or m > 12:
                errors["birthday"] = "Month must be between 1 and 12."
            elif d < 1 or d > DAYS_IN_MONTH[m]:
                errors["birthday"] = f"Day must be between 1 and {DAYS_IN_MONTH[m]} for that month."
            else:
                cleaned["birthday_month"] = m
                cleaned["birthday_day"] = d

    if "email" in p or not partial:
        email = (p.get("email") or "").strip().lower()
        if not email:
            errors["email"] = "Email is required."
        elif not EMAIL_RE.match(email):
            errors["email"] = "That doesn't look like a valid email address."
        else:
            cleaned["email"] = email

    return cleaned, errors


def row_to_dict(row):
    return {
        "id": row["id"],
        "name": row["name"],
        "role": row["role"],
        "birthday_month": row["birthday_month"],
        "birthday_day": row["birthday_day"],
        "email": row["email"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.get("/api/health")
def health():
    return jsonify({"status": "ok", "time": datetime.utcnow().isoformat() + "Z",
                    "db": "postgresql" if _IS_PG else "sqlite"})


@app.get("/api/members")
def list_members():
    with get_db() as db:
        rows = db.execute(
            f"SELECT * FROM members WHERE deleted_at IS NULL ORDER BY {_NAME_ORDER}"
        ).fetchall()
    return jsonify([row_to_dict(r) for r in rows])


@app.post("/api/members")
def create_member():
    payload = request.get_json(silent=True) or {}
    cleaned, errors = validate_member_payload(payload, partial=False)
    if errors:
        return jsonify({"errors": errors}), 400

    confirmed = bool(payload.get("confirm_duplicate"))

    with get_db() as db:
        existing_email = db.execute(
            "SELECT id, name FROM members WHERE email = ? AND deleted_at IS NULL",
            (cleaned["email"],),
        ).fetchone()
        if existing_email:
            return (
                jsonify({
                    "errors": {
                        "email": "An entry with this email already exists. Use the Edit option below to update your details.",
                    }
                }),
                409,
            )

        if not confirmed:
            soft = db.execute(
                """SELECT id, name, email FROM members
                   WHERE LOWER(name) = LOWER(?)
                     AND birthday_month = ?
                     AND birthday_day   = ?
                     AND deleted_at IS NULL""",
                (cleaned["name"], cleaned["birthday_month"], cleaned["birthday_day"]),
            ).fetchone()
            if soft:
                return (
                    jsonify({
                        "warning": {
                            "type": "soft_duplicate",
                            "message": "A person with this name and birthday already exists. Are you sure this is a different person?",
                            "existing": {"name": soft["name"], "email": soft["email"]},
                        }
                    }),
                    409,
                )

        now = datetime.utcnow().isoformat() + "Z"
        new_id = db.insert_get_id(
            """INSERT INTO members (name, role, birthday_month, birthday_day, email, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (cleaned["name"], cleaned["role"], cleaned["birthday_month"],
             cleaned["birthday_day"], cleaned["email"], now, now),
        )
        row = db.execute("SELECT * FROM members WHERE id = ?", (new_id,)).fetchone()

    write_backup()
    return jsonify(row_to_dict(row)), 201


@app.put("/api/members/<int:member_id>")
def update_member(member_id):
    payload = request.get_json(silent=True) or {}
    cleaned, errors = validate_member_payload(payload, partial=False)
    if errors:
        return jsonify({"errors": errors}), 400

    with get_db() as db:
        existing = db.execute(
            "SELECT * FROM members WHERE id = ? AND deleted_at IS NULL", (member_id,)
        ).fetchone()
        if not existing:
            return jsonify({"errors": {"_": "Member not found."}}), 404

        conflict = db.execute(
            "SELECT id FROM members WHERE email = ? AND id != ? AND deleted_at IS NULL",
            (cleaned["email"], member_id),
        ).fetchone()
        if conflict:
            return jsonify({"errors": {"email": "Another member already uses this email."}}), 409

        now = datetime.utcnow().isoformat() + "Z"
        db.execute(
            """UPDATE members
                  SET name = ?, role = ?, birthday_month = ?, birthday_day = ?, email = ?, updated_at = ?
                WHERE id = ?""",
            (cleaned["name"], cleaned["role"], cleaned["birthday_month"],
             cleaned["birthday_day"], cleaned["email"], now, member_id),
        )
        row = db.execute("SELECT * FROM members WHERE id = ?", (member_id,)).fetchone()

    write_backup()
    return jsonify(row_to_dict(row))


@app.delete("/api/members/<int:member_id>")
def delete_member(member_id):
    with get_db() as db:
        existing = db.execute(
            "SELECT id FROM members WHERE id = ? AND deleted_at IS NULL", (member_id,)
        ).fetchone()
        if not existing:
            return jsonify({"errors": {"_": "Member not found."}}), 404
        now = datetime.utcnow().isoformat() + "Z"
        db.execute("UPDATE members SET deleted_at = ? WHERE id = ?", (now, member_id))

    write_backup()
    return jsonify({"deleted": True, "id": member_id})


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
@app.get("/api/dashboard")
def dashboard():
    today = date.today()
    is_leap = (today.year % 4 == 0 and today.year % 100 != 0) or (today.year % 400 == 0)

    with get_db() as db:
        rows = db.execute("SELECT * FROM members WHERE deleted_at IS NULL").fetchall()

    todays = []
    upcoming = []

    for r in rows:
        m = r["birthday_month"]
        d = r["birthday_day"]
        observed_m, observed_d = m, d
        leap_note = False
        if m == 2 and d == 29 and not is_leap:
            observed_m, observed_d = 2, 28
            leap_note = True

        if observed_m == today.month and observed_d == today.day:
            entry = row_to_dict(r)
            entry["leap_note"] = leap_note
            todays.append(entry)
            continue

        try:
            this_year_birthday = date(today.year, observed_m, observed_d)
        except ValueError:
            continue
        if this_year_birthday < today:
            try:
                this_year_birthday = date(today.year + 1, observed_m, observed_d)
            except ValueError:
                continue
        delta = (this_year_birthday - today).days
        if 1 <= delta <= 7:
            entry = row_to_dict(r)
            entry["days_until"] = delta
            entry["next_date_iso"] = this_year_birthday.isoformat()
            entry["leap_note"] = leap_note
            upcoming.append(entry)

    upcoming.sort(key=lambda e: e["days_until"])
    todays.sort(key=lambda e: e["name"].lower())

    return jsonify({
        "today": today.isoformat(),
        "todays_birthdays": todays,
        "upcoming": upcoming,
    })


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------
def get_active_rows():
    order = f"birthday_month, birthday_day, {_NAME_ORDER}"
    with get_db() as db:
        return db.execute(
            f"SELECT * FROM members WHERE deleted_at IS NULL ORDER BY {order}"
        ).fetchall()


@app.get("/api/export/csv")
def export_csv():
    rows = get_active_rows()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Name", "Role", "Birthday Month", "Birthday Day", "Email", "Created At", "Updated At"])
    for r in rows:
        writer.writerow([r["id"], r["name"], r["role"], r["birthday_month"], r["birthday_day"],
                         r["email"], r["created_at"], r["updated_at"]])
    data = output.getvalue().encode("utf-8")
    buf = io.BytesIO(data)
    buf.seek(0)
    return send_file(buf, mimetype="text/csv", as_attachment=True,
                     download_name=f"birthdays_{datetime.utcnow().strftime('%Y%m%d')}.csv")


@app.get("/api/export/xlsx")
def export_xlsx():
    rows = get_active_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Birthdays"
    headers = ["ID", "Name", "Role", "Birthday Month", "Birthday Day", "Email", "Created At", "Updated At"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for r in rows:
        ws.append([r["id"], r["name"], r["role"], r["birthday_month"], r["birthday_day"],
                   r["email"], r["created_at"], r["updated_at"]])
    for col_idx, _ in enumerate(headers, start=1):
        max_len = max(
            (len(str(ws.cell(row=row, column=col_idx).value or "")) for row in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=f"birthdays_{datetime.utcnow().strftime('%Y%m%d')}.xlsx")


@app.get("/api/backup/latest")
def latest_backup_info():
    if _IS_PG:
        return jsonify({
            "exists": False,
            "note": "File backups disabled in PostgreSQL mode — data persists in the cloud database.",
        })
    if not os.path.exists(LATEST_BACKUP):
        return jsonify({"exists": False})
    stat = os.stat(LATEST_BACKUP)
    return jsonify({
        "exists": True,
        "modified_at": datetime.utcfromtimestamp(stat.st_mtime).isoformat() + "Z",
        "size_bytes": stat.st_size,
    })


# ---------------------------------------------------------------------------
# Bootstrap
# ---------------------------------------------------------------------------
init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    app.run(host="0.0.0.0", port=port, debug=False)
